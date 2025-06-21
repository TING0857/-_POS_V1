# inventory.py｜良級懸賞 POS 系統 — 庫存管理功能
import os, sys, subprocess, json, tkinter as tk, csv
from tkinter import ttk, filedialog, messagebox
import webbrowser
from datetime import datetime
import logs
import receive

# 確保 openpyxl 安裝
try:
    from openpyxl import load_workbook
except ImportError:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    from openpyxl import load_workbook

# 資料儲存路徑
DATA_DIR = os.path.join(os.path.dirname(__file__), 'logs')
os.makedirs(DATA_DIR, exist_ok=True)
INVENTORY_FILE = os.path.join(DATA_DIR, 'inventory.json')

class InventoryApp:
    def __init__(self, root):
        self.root = root
        self.root.title('良級懸賞 POS 系統')
        self.root.geometry('1000x600')
        self.data = []
        self.load_data()

        # inline edit 狀態
        self.editing = False
        self.inline_entry = None
        self.current_row = None
        self.current_col = 0

        self.build_ui()
        self.refresh_table()

    def load_data(self):
        if os.path.exists(INVENTORY_FILE):
            with open(INVENTORY_FILE, 'r', encoding='utf-8') as f:
                self.data = json.load(f)
        else:
            self.data = []

    def save_data(self):
        with open(INVENTORY_FILE, 'w', encoding='utf-8') as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)

    def build_ui(self):
        # 建立 Notebook 分頁
        notebook = ttk.Notebook(self.root)
        notebook.pack(fill='both', expand=True)

        # —— 分頁 1：庫存管理
        frame1 = ttk.Frame(notebook)
        notebook.add(frame1, text='庫存管理')

        # 上方工具列
        top = ttk.Frame(frame1)
        top.pack(fill='x', padx=10, pady=5)
        self.search_var = tk.StringVar()
        entry = ttk.Entry(top, textvariable=self.search_var)
        entry.pack(side='left', fill='x', expand=True)
        entry.insert(0, '搜尋')
        entry.bind('<FocusIn>', lambda e: self.search_var.set('') if self.search_var.get()=='搜尋' else None)
        entry.bind('<FocusOut>', lambda e: self.search_var.set('搜尋') if not self.search_var.get() else None)
        entry.bind('<Return>', lambda e: self.refresh_table())
        ttk.Button(top, text='搜尋',   command=self.refresh_table).pack(side='left', padx=5)
        ttk.Button(top, text='新增商品', command=self.open_add_dialog).pack(side='left', padx=5)
        ttk.Button(top, text='批次匯入', command=self.batch_import).pack(side='left', padx=5)
        ttk.Button(top, text='匯出',   command=self.export_data).pack(side='left', padx=5)
        ttk.Button(top, text='關班',   command=self.on_close_shift).pack(side='left', padx=5)

        # 欄位設定
        self.cols = [
            '廠商','關鍵字IP','編碼','商品名稱','數量','成本','點數價',
            '20洞價格','40洞價格','60洞價格','80洞價格','備註','商品連結'
        ]
        self.tree = ttk.Treeview(
            frame1, columns=self.cols, show='headings', selectmode='extended'
        )
        for c in self.cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=100, anchor='center')
        self.tree.column('商品名稱', width=200)
        self.tree.column('備註', width=120)
        self.tree.column('商品連結', width=200)
        self.tree.pack(fill='both', expand=True, padx=10, pady=5)

        vsb = ttk.Scrollbar(frame1, orient='vertical', command=self.tree.yview)
        vsb.place(relx=1.0, rely=0, relheight=1.0, anchor='ne')
        self.tree.configure(yscrollcommand=vsb.set)

        # 綁定事件
        self.tree.bind('<Button-1>',    self.on_single_click)
        self.tree.bind('<Double-1>',    self.on_double_click)
        self.tree.bind('<Return>',      self.on_enter_key)
        self.tree.bind('<Button-3>',    self.show_context_menu)
        self.tree.bind('<Delete>',      self.delete_selected)
        self.tree.bind('<B1-Motion>',   self.on_drag_select)
        for k in ('<Up>','<Down>','<Left>','<Right>'):
            self.tree.bind(k, lambda e: None)

        # 右鍵選單
        self.menu = tk.Menu(self.root, tearoff=0)
        self.menu.add_command(label='編輯',    command=self.edit_dialog)
        self.menu.add_command(label='刪除',    command=self.delete_item)
        self.menu.add_command(label='抽賞結帳', command=self.checkout_item)

        # —— 分頁 2：交易紀錄
        frame2 = ttk.Frame(notebook)
        notebook.add(frame2, text='交易紀錄')
        self.logs_frame = logs.LogsFrame(frame2)
        self.logs_frame.pack(fill='both', expand=True)

        # —— 分頁 3：商品領取紀錄
        frame3 = ttk.Frame(notebook)
        notebook.add(frame3, text='商品領取紀錄')
        self.receive_frame = receive.ReceiveFrame(frame3)
        self.receive_frame.pack(fill='both', expand=True)

        # 綁定結帳完成事件，自動刷新領取紀錄
        self.root.bind('<<TransactionDone>>', lambda e: (
            self.receive_frame.load_data(),
            self.receive_frame.refresh()
        ))

    def on_drag_select(self, event):
        row = self.tree.identify_row(event.y)
        if row:
            self.tree.selection_add(row)

    def refresh_table(self):
        kw = self.search_var.get().strip()
        if kw == '搜尋': kw = ''
        self.tree.delete(*self.tree.get_children())
        for idx, item in enumerate(self.data):
            text = json.dumps(item, ensure_ascii=False)
            if kw.lower() in text.lower():
                vals = [item.get(c,'') for c in self.cols]
                self.tree.insert('', 'end', iid=str(idx), values=vals)

    def show_context_menu(self, event):
        iid = self.tree.identify_row(event.y)
        if iid:
            self.tree.selection_set(iid)
            self.menu.tk_popup(event.x_root, event.y_root)

    def on_single_click(self, event):
        row = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if row and col == f"#{len(self.cols)}":
            url = self.tree.set(row, col)
            if url:
                webbrowser.open_new(url)

    def on_double_click(self, event):
        row = self.tree.identify_row(event.y)
        if row:
            self.tree.selection_set(row)
            self.checkout_item()

    def on_enter_key(self, event):
        if not self.editing:
            row = self.tree.focus()
            if row:
                col = f"#{self.current_col+1}"
                self.start_inline_edit(row, col)
        else:
            self.finish_inline_edit()

    def _navigate_cell(self, direction):
        if not self.editing:
            return
        self.finish_inline_edit()
        row = int(self.current_row)
        col = self.current_col
        if direction == 'Left':  col = max(0, col-1)
        if direction == 'Right': col = min(len(self.cols)-1, col+1)
        if direction == 'Up':    row = max(0, row-1)
        if direction == 'Down':  row = min(len(self.data)-1, row+1)
        self.current_col = col
        self.start_inline_edit(str(row), f"#{col+1}")

    def start_inline_edit(self, row, col):
        x, y, w, h = self.tree.bbox(row, col)
        val = self.tree.set(row, col)
        entry = ttk.Entry(self.tree)
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, val)
        entry.focus()
        entry.bind('<Return>', self._save_inline)
        entry.bind('<FocusOut>', lambda e: self.finish_inline_edit())
        entry.bind('<Left>',  lambda e: self._navigate_cell('Left'))
        entry.bind('<Right>', lambda e: self._navigate_cell('Right'))
        entry.bind('<Up>',    lambda e: self._navigate_cell('Up'))
        entry.bind('<Down>',  lambda e: self._navigate_cell('Down'))
        self.inline_entry = entry
        self.editing = True
        self.current_row = row
        self.current_col = int(col.replace('#','')) - 1

    def _save_inline(self, event):
        self.finish_inline_edit()

    def finish_inline_edit(self):
        new = self.inline_entry.get()
        row, col = self.current_row, f"#{self.current_col+1}"
        self.inline_entry.destroy()
        self.inline_entry = None
        self.editing = False
        self.tree.set(row, col, new)
        self.data[int(row)][self.cols[self.current_col]] = new
        self.save_data()

    def open_add_dialog(self):
        dlg = tk.Toplevel(self.root)
        dlg.title('新增商品')
        entries = {}
        for i, c in enumerate(self.cols):
            ttk.Label(dlg, text=c).grid(row=i, column=0, padx=5, pady=2, sticky='e')
            ent = ttk.Entry(dlg)
            ent.grid(row=i, column=1, padx=5, pady=2)
            entries[c] = ent

        def calc(e=None):
            try:
                cost = float(entries['成本'].get())
                dot  = int(cost * 1.2)
                entries['點數價'].delete(0,'end'); entries['點數價'].insert(0,str(dot))
                p20 = int(cost / 7)
                entries['20洞價格'].delete(0,'end'); entries['20洞價格'].insert(0,str(p20))
                for col_name, div in [('40洞價格',2),('60洞價格',3),('80洞價格',4)]:
                    val = int((p20+100)/div)
                    entries[col_name].delete(0,'end'); entries[col_name].insert(0,str(val))
            except:
                pass

        entries['成本'].bind('<FocusOut>', calc)
        btn = ttk.Button(dlg, text='確定', command=lambda: self._add_and_close(entries, dlg))
        btn.grid(row=len(self.cols), column=0, columnspan=2, pady=10)

    def _add_and_close(self, entries, dlg):
        item = {c: entries[c].get().strip() for c in self.cols}
        self.data.append(item)
        self.save_data()
        self.refresh_table()
        dlg.destroy()

    def edit_dialog(self):
        sel = self.tree.selection()
        if not sel:
            return
        idx = int(sel[0])
        item = self.data[idx]
        dlg = tk.Toplevel(self.root)
        dlg.title('編輯商品')
        entries = {}
        for i, c in enumerate(self.cols):
            ttk.Label(dlg, text=c).grid(row=i, column=0, padx=5, pady=2, sticky='e')
            ent = ttk.Entry(dlg)
            ent.insert(0, item.get(c, ''))
            ent.grid(row=i, column=1, padx=5, pady=2)
            entries[c] = ent

        btn = ttk.Button(dlg, text='儲存', command=lambda: self._save_and_close(idx, entries, dlg))
        btn.grid(row=len(self.cols), column=0, columnspan=2, pady=10)

    def _save_and_close(self, idx, entries, dlg):
        for c in self.cols:
            self.data[idx][c] = entries[c].get().strip()
        self.save_data()
        self.refresh_table()
        dlg.destroy()

    def delete_item(self):
        sels = self.tree.selection()
        if not sels:
            return
        idx = int(sels[0])
        if messagebox.askyesno('刪除商品', '確定要刪除此商品？'):
            self.data.pop(idx)
            self.save_data()
            self.refresh_table()

    def delete_selected(self, event=None):
        sels = self.tree.selection()
        if not sels:
            return
        count = len(sels)
        if not messagebox.askyesno('刪除商品', f'確定要刪除選取的 {count} 筆商品？'):
            return
        for i in sorted([int(x) for x in sels], reverse=True):
            self.data.pop(i)
        self.save_data()
        self.refresh_table()

    def checkout_item(self):
        # 取消自動填“已被抽走”，保留原有結帳流程
        sels = self.tree.selection()
        if not sels:
            return
        idx = int(sels[0])
        popup = tk.Toplevel(self.root)
        import checkout
        checkout.CheckoutApp(popup, idx)
        popup.wait_window()

        # 結帳完成後，重新載入所有
        self.load_data()
        self.refresh_table()
        self.logs_frame.load_all_logs()
        self.logs_frame.refresh_logs()
        self.receive_frame.load_data()
        self.receive_frame.refresh()

    def batch_import(self):
        path = filedialog.askopenfilename(
            title='選擇 Excel 檔', filetypes=[('Excel','*.xlsx')]
        )
        if not path:
            return
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        imported = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            code = row[0] or ''
            name = row[1] or ''
            link = row[2] or ''
            kw = (row[3] or '').split()[0] if row[3] else ''
            cost_val = row[9] or row[8] or 0
            cost = int(cost_val)
            dot = int(cost * 1.2)
            p20 = int(cost / 7)
            p40 = int((p20+100)/2)
            p60 = int((p20+100)/3)
            p80 = int((p20+100)/4)
            raw_qty = row[11]
            qty = str(int(raw_qty)) if isinstance(raw_qty, (int, float)) else str(raw_qty or '')
            item = {
                '廠商':'良級懸賞','關鍵字IP':kw,'編碼':code,'商品名稱':name,
                '數量':qty,'成本':str(cost),'點數價':str(dot),
                '20洞價格':str(p20),'40洞價格':str(p40),
                '60洞價格':str(p60),'80洞價格':str(p80),
                '備註':'','商品連結':link
            }
            self.data.append(item)
            imported += 1
        self.save_data()
        self.refresh_table()
        messagebox.showinfo('匯入完成', f'已匯入 {imported} 筆資料')

    def export_data(self):
        path = filedialog.asksaveasfilename(
            defaultextension='.json', filetypes=[('JSON','*.json')]
        )
        if not path:
            return
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)
        messagebox.showinfo('匯出完成', f'已匯出到 {path}')

    def on_close_shift(self):
        if not messagebox.askyesno('關班確認', '關班後將匯出今日交易與領取報表並關閉系統，是否繼續？'):
            return
        today = datetime.now().strftime('%Y-%m-%d')
        closing_dir = os.path.join(DATA_DIR, 'closing')
        os.makedirs(closing_dir, exist_ok=True)

        # 匯出交易日誌
        logs_file = logs.LOG_FILE
        logs_csv  = os.path.join(closing_dir, f'logs_{today}.csv')
        with open(logs_file, 'r', encoding='utf-8') as f_in, \
             open(logs_csv, 'w', newline='', encoding='utf-8') as f_out:
            writer = csv.writer(f_out)
            writer.writerow(['time','branch','staff','member','item','hole','抽數','大賞','小賞','cash','transfer','points','total','reason'])
            for line in f_in:
                try:
                    rec = json.loads(line)
                except:
                    continue
                if rec.get('time','')[:10] == today:
                    writer.writerow([rec.get(k,'') for k in [
                        'time','branch','staff','member','item','hole','抽數','大賞','小賞','cash','transfer','points','total','reason'
                    ]])

        # 匯出領取紀錄
        recv_file = receive.RECEIVE_FILE
        recv_csv  = os.path.join(closing_dir, f'receive_{today}.csv')
        with open(recv_file, 'r', encoding='utf-8') as f_in, \
             open(recv_csv, 'w', newline='', encoding='utf-8') as f_out:
            writer = csv.writer(f_out)
            writer.writerow(['日期','member','item','qty','expire','free','reason','已領取'])
            text = f_in.read().strip()
            all_data = []
            if text.startswith('['):
                try:
                    all_data = json.loads(text)
                except:
                    all_data = []
            else:
                for line in text.splitlines():
                    try:
                        all_data.append(json.loads(line))
                    except:
                        continue
            for rec in all_data:
                if rec.get('日期','')[:10] == today:
                    writer.writerow([
                        rec.get('日期',''),
                        rec.get('member',''),
                        rec.get('item',''),
                        rec.get('inventory_qty', rec.get('qty','')),
                        rec.get('expire', rec.get('到期日','')),
                        rec.get('free',''),
                        rec.get('reason',''),
                        '✔' if rec.get('已領取') else ''
                    ])

        messagebox.showinfo('完成', f'已匯出今日報表至 {closing_dir}，系統即將關閉。')
        self.root.destroy()

if __name__ == '__main__':
    root = tk.Tk()
    app = InventoryApp(root)
    root.mainloop()
